from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()
sheet = wb.active
sheet.title = "Trending Stocks"

# Title row
sheet['A1'] = '화제 종목 TOP 5 - 2025.12.05'
sheet['A1'].font = Font(size=14, bold=True)
sheet['A1'].alignment = Alignment(horizontal='center')
sheet.merge_cells('A1:E1')

# Headers
headers = ['순위', '티커', '종목명', '주가', '등락률']
for col, header in enumerate(headers, 1):
    cell = sheet.cell(row=2, column=col)
    cell.value = header
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center')

# Data
data = [
    [1, 'TSLA', 'Tesla', 385, 0.087],
    [2, 'NVDA', 'NVIDIA', 142, 0.052],
    [3, 'AAPL', 'Apple', 195, 0.021]
]

# Border style
thin = Side(border_style='thin', color='000000')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for row_idx, row_data in enumerate(data, 3):
    for col_idx, value in enumerate(row_data, 1):
        cell = sheet.cell(row=row_idx, column=col_idx)
        cell.value = value
        cell.border = border

        # 순위 - center alignment
        if col_idx == 1:
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # 주가 - currency format
        if col_idx == 4:
            cell.number_format = '$#,##0'
            cell.alignment = Alignment(horizontal='right', vertical='center')

        # 등락률 - percentage with color
        if col_idx == 5:
            cell.number_format = '0.0%'
            cell.alignment = Alignment(horizontal='right', vertical='center')
            if value > 0:
                cell.font = Font(color='00B050', bold=True)  # Green
            elif value < 0:
                cell.font = Font(color='FF0000', bold=True)  # Red

# Add borders to headers
for col in range(1, 6):
    sheet.cell(row=2, column=col).border = border

# Column widths
sheet.column_dimensions['A'].width = 8
sheet.column_dimensions['B'].width = 10
sheet.column_dimensions['C'].width = 15
sheet.column_dimensions['D'].width = 12
sheet.column_dimensions['E'].width = 12

# Row height
sheet.row_dimensions[1].height = 25
sheet.row_dimensions[2].height = 20

# 동적 경로 생성
from pathlib import Path
import os
from datetime import datetime

# 프로젝트 루트 찾기 (스크립트 위치 기준)
script_dir = Path(__file__).parent
project_root = script_dir.parent.parent.parent
output_dir = project_root / 'backend' / 'output' / 'data'
output_dir.mkdir(parents=True, exist_ok=True)

# 날짜 기반 파일명 생성
date_str = datetime.now().strftime('%Y-%m-%d')
output_path = output_dir / f'trending_{date_str}.xlsx'

wb.save(str(output_path))
print(f'Excel file created successfully: {output_path}')
