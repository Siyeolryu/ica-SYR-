"""
Excel 파일 생성 모듈
화제 종목 데이터를 Excel 파일로 생성
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path
from datetime import datetime
from typing import List, Dict
import logging

logger = logging.getLogger(__name__)


def create_trending_stocks_excel(
    stocks: List[Dict],
    output_dir: str = None,
    filename: str = None
) -> str:
    """
    화제 종목 데이터를 Excel 파일로 생성
    
    Args:
        stocks: 종목 데이터 리스트
            각 항목은 다음 키를 포함해야 함:
            - symbol: 종목 심볼 (예: 'AAPL')
            - name: 종목명 (예: 'Apple Inc.')
            - price: 현재가
            - change_percent: 등락률 (소수점, 예: 0.05 = 5%)
        output_dir: 출력 디렉토리 (기본값: backend/output/data)
        filename: 파일명 (기본값: trending_YYYY-MM-DD.xlsx)
    
    Returns:
        생성된 Excel 파일 경로
    """
    try:
        # 출력 경로 설정
        if output_dir is None:
            script_dir = Path(__file__).parent
            output_dir = script_dir / 'output' / 'data'
        else:
            output_dir = Path(output_dir)
        
        output_dir.mkdir(parents=True, exist_ok=True)
        
        if filename is None:
            date_str = datetime.now().strftime('%Y-%m-%d')
            filename = f'trending_{date_str}.xlsx'
        
        output_path = output_dir / filename
        
        # Workbook 생성
        wb = Workbook()
        sheet = wb.active
        sheet.title = "Trending Stocks"
        
        # 제목 행
        title = f'화제 종목 TOP {len(stocks)} - {datetime.now().strftime("%Y.%m.%d")}'
        sheet['A1'] = title
        sheet['A1'].font = Font(size=14, bold=True)
        sheet['A1'].alignment = Alignment(horizontal='center')
        sheet.merge_cells(f'A1:E1')
        
        # 헤더
        headers = ['순위', '티커', '종목명', '주가', '등락률']
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=2, column=col)
            cell.value = header
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 테두리 스타일
        thin = Side(border_style='thin', color='000000')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        
        # 데이터 입력
        for idx, stock in enumerate(stocks, 1):
            row_idx = idx + 2  # 헤더 다음 행부터
            
            # 데이터 추출
            symbol = stock.get('symbol', 'N/A')
            name = stock.get('name', 'N/A')
            price = stock.get('price', 0)
            change_percent = stock.get('change_percent', 0)
            
            # 데이터 입력
            sheet.cell(row=row_idx, column=1).value = idx  # 순위
            sheet.cell(row=row_idx, column=2).value = symbol  # 티커
            sheet.cell(row=row_idx, column=3).value = name  # 종목명
            sheet.cell(row=row_idx, column=4).value = price  # 주가
            sheet.cell(row=row_idx, column=5).value = change_percent  # 등락률
            
            # 셀 스타일 적용
            for col_idx in range(1, 6):
                cell = sheet.cell(row=row_idx, column=col_idx)
                cell.border = border
                
                # 순위 - 중앙 정렬
                if col_idx == 1:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # 주가 - 통화 형식, 우측 정렬
                if col_idx == 4:
                    cell.number_format = '$#,##0.00'
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                
                # 등락률 - 퍼센트 형식, 색상 적용
                if col_idx == 5:
                    cell.number_format = '0.00%'
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                    if change_percent > 0:
                        cell.font = Font(color='00B050', bold=True)  # 녹색
                    elif change_percent < 0:
                        cell.font = Font(color='FF0000', bold=True)  # 빨간색
        
        # 헤더 테두리 추가
        for col in range(1, 6):
            sheet.cell(row=2, column=col).border = border
        
        # 열 너비 설정
        sheet.column_dimensions['A'].width = 8
        sheet.column_dimensions['B'].width = 10
        sheet.column_dimensions['C'].width = 20
        sheet.column_dimensions['D'].width = 12
        sheet.column_dimensions['E'].width = 12
        
        # 행 높이 설정
        sheet.row_dimensions[1].height = 25
        sheet.row_dimensions[2].height = 20
        
        # 파일 저장
        wb.save(str(output_path))
        logger.info(f"Excel 파일 생성 완료: {output_path}")
        
        return str(output_path)
        
    except Exception as e:
        logger.error(f"Excel 파일 생성 실패: {str(e)}")
        raise


if __name__ == '__main__':
    # 테스트 데이터
    test_stocks = [
        {
            'symbol': 'TSLA',
            'name': 'Tesla, Inc.',
            'price': 385.50,
            'change_percent': 0.087
        },
        {
            'symbol': 'NVDA',
            'name': 'NVIDIA Corporation',
            'price': 142.30,
            'change_percent': 0.052
        },
        {
            'symbol': 'AAPL',
            'name': 'Apple Inc.',
            'price': 195.20,
            'change_percent': 0.021
        }
    ]
    
    output_path = create_trending_stocks_excel(test_stocks)
    print(f"Excel 파일 생성 완료: {output_path}")

